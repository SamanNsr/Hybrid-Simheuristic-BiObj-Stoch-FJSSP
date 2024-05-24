
import time
import pandas as pd
import numpy as np
from typing import Union, List
import plotly.express as px
from openpyxl import load_workbook

from .metaheuristic import Problem, IntegerVar, Solution


class SimMOFJSP(Problem):
    def __init__(self, file_path, number_of_objectives=2, sim=True, uncertain_factor=0.1, initial_solution=[], **kwargs):
        super(SimMOFJSP, self).__init__()
        workbook = load_workbook(file_path)
        info_sheet = workbook.active
        num_jobs = info_sheet.cell(row=1, column=2).value
        num_machines = info_sheet.cell(row=2, column=2).value
        self.uncertain_factor = uncertain_factor
        self.initial_solution = initial_solution

        self.sim = sim
        self.nov = num_machines * num_jobs
        self.lower_bound = self.nov * [0]
        self.upper_bound = self.nov * [num_jobs - 1]
        IntegerVar.lower_bound = self.lower_bound
        IntegerVar.upper_bound = self.upper_bound

        self.objective_directions = [self.MINIMIZE] * number_of_objectives
        self.objective_labels = [
            f"$ f_{{ {i} }} $" for i in range(number_of_objectives)]

        self.num_jobs = num_jobs
        self.num_machines = num_machines
        processing_times_data = pd.read_excel(
            file_path, sheet_name="Processing Time", index_col=[0])
        machine_sequences_data = pd.read_excel(
            file_path, sheet_name="Machines Sequence", index_col=[0])
        due_dates_data = pd.read_excel(
            file_path, sheet_name="Priority and Due date", index_col=[0])
        self.processing_times = [list(map(int, processing_times_data.iloc[i]))
                                 for i in range(self.num_jobs)]
        self.machine_sequences = [list(map(int, machine_sequences_data.iloc[i]))
                                  for i in range(self.num_jobs)]
        self.due_dates = [
            list(due_dates_data.iloc[i]) for i in range(self.num_jobs)]
        self.start_time = time.time()
        self.iteration_counter = 0

    def evaluate(self, solution: Solution) -> Solution:
        x = solution.variables.copy()
        if isinstance(x, np.ndarray):
            x = x.tolist()
        x = self.correct_solution(x)
        if self.sim:
            twte, makespan, pt = self.monte_carlo_simulation(x, 10)
        else:
            twte, makespan = self.calculate_objective(x, self.pt)
        solution.objectives = [twte, makespan]
        return solution

    def normal_processing_time(self, pt):
        pt = np.array(pt)
        return np.random.normal(pt, pt/10)

    def log_normal_processing_time(self, pt):
        pt = np.array(pt)
        var = self.uncertain_factor * pt
        sigma = np.abs(np.sqrt(np.log(1 + var/pt**2)))
        return np.random.lognormal(np.log(pt), sigma)

    def monte_carlo_simulation(self, x, num_iter=100):
        total_twte, total_makespan = 0, 0
        for i in range(num_iter):
            stochastic_pts = self.log_normal_processing_time(self.pt)
            twte, makespan = self.calculate_objective(x, stochastic_pts)
            total_twte += twte
            total_makespan += makespan
        twte = total_twte / num_iter
        makespan = total_makespan / num_iter
        return twte, makespan, stochastic_pts

    def exponential_processing_time(self, pt):
        return np.random.exponential(pt)

    def calculate_objective(self, sequence, processing_time):
        job_indices = [job_index for job_index in range(self.num_jobs)]
        process_step_counts = {index: 0 for index in job_indices}
        job_completion_times = {index: 0 for index in job_indices}
        machine_indices = [index + 1 for index in range(self.num_machines)]
        machine_completion_times = {index: 0 for index in machine_indices}
        delay_record = {}

        for job in sequence:
            job = int(job)
            processing_duration = int(
                processing_time[job][process_step_counts[job]])
            machine_assigned = int(
                self.machine_sequences[job][process_step_counts[job]])
            job_completion_times[job] += processing_duration
            machine_completion_times[machine_assigned] += processing_duration

            if machine_completion_times[machine_assigned] < job_completion_times[job]:
                machine_completion_times[machine_assigned] = job_completion_times[job]
            elif machine_completion_times[machine_assigned] > job_completion_times[job]:
                job_completion_times[job] = machine_completion_times[machine_assigned]

            process_step_counts[job] += 1

        for job_index in job_indices:
            if job_completion_times[job_index] > self.due_dates[job_index][1]:
                tardiness = job_completion_times[job_index] - \
                    self.due_dates[job_index][1]
                earliness = 0
                delay_record[job_index] = [earliness, tardiness]
            elif job_completion_times[job_index] < self.due_dates[job_index][1]:
                tardiness = 0
                earliness = self.due_dates[job_index][1] - \
                    job_completion_times[job_index]
                delay_record[job_index] = [earliness, tardiness]
            else:
                tardiness = 0
                earliness = 0
                delay_record[job_index] = [earliness, tardiness]

        total_weighted_tardiness = sum((1 / self.due_dates[job_index][0]) * delay_record[job_index][0] +
                                       self.due_dates[job_index][0] * delay_record[job_index][1] for job_index in job_indices)
        makespan = max(job_completion_times.values())
        return total_weighted_tardiness, makespan

    def correct_solution(self, job_sequence):
        adjusted_sequence = np.array(job_sequence).tolist() if isinstance(
            job_sequence, np.ndarray) else job_sequence

        job_counts = {}
        over_assigned_jobs, under_assigned_jobs = [], []
        for job_id in range(self.num_jobs):
            if job_id in adjusted_sequence:
                count = adjusted_sequence.count(job_id)
                first_position = adjusted_sequence.index(job_id)
                job_counts[job_id] = [count, first_position]
            else:
                job_counts[job_id] = [0, 0]
            if count > self.num_machines:
                over_assigned_jobs.append(job_id)
            elif count < self.num_machines:
                under_assigned_jobs.append(job_id)

        for job_id in over_assigned_jobs:
            while job_counts[job_id][0] > self.num_machines:
                for under_job in under_assigned_jobs:
                    if job_counts[under_job][0] < self.num_machines:
                        adjusted_sequence[job_counts[job_id][1]] = under_job
                        job_counts[job_id][1] = adjusted_sequence.index(job_id)
                        job_counts[job_id][0] -= 1
                        job_counts[under_job][0] += 1
                    if job_counts[job_id][0] == self.num_machines:
                        break

        return np.array(adjusted_sequence)

    def create_solution(self, encoded: bool = True) -> Union[List, np.ndarray]:
        new_solution = Solution(
            self.lower_bound, self.upper_bound, self.number_of_objectives(
            ), self.number_of_constraints()
        )
        if len(self.initial_solution) > self.counter:
            new_solution.variables = self.initial_solution[self.counter]
            self.counter += 1
            return new_solution
        self.counter += 1
        unique_numbers = np.random.choice(
            np.arange(self.num_jobs), self.num_mc, replace=False)
        result_array = np.repeat(unique_numbers, self.num_jobs)
        np.random.shuffle(result_array)
        new_solution.variables = result_array.tolist()
        return new_solution
